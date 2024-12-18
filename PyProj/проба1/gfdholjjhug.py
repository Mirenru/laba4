import telebot
from telebot import types
import requests
import json
import logging
import openpyxl
from openpyxl import load_workbook
API_TOKEN = '7591208613:AAElKLMG97nG8O9PLXQ2GVOFwIj-KDas5BA'
spell_check_enabled = True
bot = telebot.TeleBot(API_TOKEN)
file = "bd.xlsx"
bd = openpyxl.open(file, read_only=False)
wb = load_workbook(file)
ws = wb['bd']
sheet = bd.active
checkflag = True
class SpellChecker:
    def __init__(self, api_url="https://speller.yandex.net/services/spellservice.json/checkText"):
        self.api_url = api_url

    def check_spelling(self, text):
        try:
            
            params = {
                "text": text

            }
            response = requests.post(self.api_url, params=params, timeout=50)

            
            logging.debug(f"Response content: {response.content[:100]}...")
            
            data = response.json()
            
            for wod in data:
                return [{wod["word"],wod["s"]}]
            else:
                return None
        except requests.RequestException as e:
            logging.error(f"Request error: {e}")
            return f"Произошла ошибка при запросе к сервису проверки орфографии: {str(e)}"
        except json.JSONDecodeError as e:
            logging.error(f"JSON decoding error: {e}")
            return f"Не удалось разобрать ответ от сервиса проверки орфографии: {str(e)}"
        except Exception as e:
            logging.error(f"Unexpected error: {e}")
            return f"Произошла непредвиденная ошибка: {str(e)}"

spell_checker = SpellChecker()


@bot.message_handler(commands=['start'])
def cmd_start(message):
    bot.reply_to(message, "Привет! Я бот для управления заметками.\n"
                        "/add - добавить заметку\n"
                        "/view - посмотреть все заметки\n"
                        "/edit - изменить последнюю заметку\n"
                        "/delete - удалить последнюю заметку\n"
                        "/help - просмот всех доступных команд\n")

@bot.message_handler(commands=['help'])
def cmd_help(message):

    bot.reply_to(message, "Привет! Напомню что я умею.\n"
                        "/add - добавить заметку\n"
                        "/view - посмотреть все заметки\n"
                        "/edit - изменить последнюю заметку\n"
                        "/delete - удалить последнюю заметку\n"
                        )



@bot.message_handler(commands=['add'])
def cmd_add(message):
    sent = bot.reply_to(message, "Напишите текст вашей заметки")
    bot.register_next_step_handler(sent, adding)


def adding(message):
    note_text = message.text

    if note_text.startswith('/'):  # Если введено начало команды
        bot.reply_to(message, "Пожалуйста, введите текст заметки, а не команду.")
        return
    # if spell_check_enabled:
    #     errors = spell_checker.check_spelling(str(note_text))
    #     if isinstance(errors, str) and errors.startswith("Произошла ошибка"):  # Если произошла ошибка
    #         bot.reply_to(message, errors)
    #         return
        
    #     if errors:
    #         error_message = ""
    #         for error in errors:
                
    #             suggestions = ", ".join(error["s"])
    #             error_message += f"В слове {error} допущена ошибка. Предлагаемые варианты: {suggestions}\n"
            
    #         bot.reply_to(message, error_message + "\nВы хотите исправить текст?")
    #         sheet.append(message.chat.id, {"note_text": note_text})
    #         sheet.delete_rows(sheet.max_row-1)
    #         return

    sheet.append([message.chat.id, note_text])

    bd.save(file)

    bot.reply_to(message, "Заметка сохранена")


@bot.message_handler(commands=['view'])
def cmd_view(message):
    k = 0
    id = message.chat.id
    bot.reply_to(message, "Ваши заметки:")
    for row in range(2, sheet.max_row+1):
        if (id == sheet[row][0].value):
            k = k+1
            bot.reply_to(message, f"{sheet[row][1].value}\n номер заметки: {k}")
        else:
            continue
            


@bot.message_handler(commands=['edit'])
def cmd_edit(message):
    sent = bot.reply_to(message, "Напишите номер заметки которую вы хотите изменит. Для просмотра своих заметок используйте команду /view.")
    bot.register_next_step_handler(sent, edding)
                
def edding(message):
    c = int(message.text)
    del_ = c
    id = message.chat.id
    for row in range(2, sheet.max_row+1):
        if (id == sheet[row][0].value):
            if (c>1):
                c = c - 1
            elif(c == 1):
                bot.reply_to(message, f"{str(sheet[del_+1][1].value)}")


                bd.save(file)
                
                sent = bot.reply_to(message, "Напишите текст вашей изменённой заметки")
                sheet.delete_rows(del_+1)
                bot.register_next_step_handler(sent, adding)
                break
                


@bot.message_handler(commands=['delete'])
def cmd_delete(message):
     sent = bot.reply_to(message, "Напишите номер заметки которую вы хотите изменит. Для просмотра своих заметок используйте команду /view.")
     bot.register_next_step_handler(sent, deleting)
     


def deleting(message):
    id = message.chat.id
    c = int(message.text)
    for row in range(2, sheet.max_row+1):
        if (id == sheet[row][0].value):
            if (c>1):
                c = c - 1
                continue
            elif(c == 1):
                c = c - 1
                sheet.delete_rows(row)
                print(row)
                bd.save(file)
                bot.reply_to(message,"Заметка удалена")
                continue
            else:
                bot.reply_to(message, "Заметка с таким номером не найдена")


@bot.message_handler(commands=['spellcheck'])
def cmd_spellcheck():
    print()

@bot.message_handler(commands=['options'])
def options(message):
    bot.reply_to(message, "Настройте бота")




bot.polling()